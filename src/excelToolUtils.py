#读取2007以后的excel的库
import openpyxl

def writeExcel(path,sheetTitle,value):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheetTitle

    # value = [["名称", "价格", "出版社", "语言"],
    #          ["如何高效读懂一本书", "22.3", "机械工业出版社", "中文"],
    #          ["暗时间", "32.4", "人民邮电出版社", "中文"],
    #          ["拆掉思维里的墙", "26.7", "机械工业出版社", "中文"]]
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    wb.save(path)
    print("写入数据成功！")


def readExcel(path):
    """
    :param path: 文件路径
    """
    # 读取文档
    wb = openpyxl.load_workbook(path)
    # 默认取第一个sheet页
    sheet = wb.worksheets[0]
    # 循环行index
    c=1
    message=""
    # 文档的信息 key=微信号  value=文本信息
    dict ={}
    for row in sheet.rows:
        # 表头不处理
        if c == 1:
            c = c + 1
            continue
        # row=行数，column=列数
        # 微信号
        weiXin = sheet.cell(row=c, column=4).value

        if weiXin is None:
            weiXin = "其他"
        if not weiXin.strip():
            print(c)
        # 微信号对应的每行文本信息
        message = sheet.cell(row=c, column=7).value
        if message is None:
            message = ""
        if weiXin in dict.keys() :
            dict[weiXin] = dict[weiXin] + message + "。"
        else:
            dict[weiXin] = message
        c = c + 1
    return dict
